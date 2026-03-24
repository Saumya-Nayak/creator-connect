"""
routes/admin/analytics_routes.py
─────────────────────────────────────────────────────────────────────────────
Admin Analytics & Reports Routes — CreatorConnect

PDF EXPORT — matches settings_pdf.js exactly:
  • A4 portrait  (210 × 297 mm)
  • Header: white left strip with logo, pink rounded bar on the right
  • Brand pink  rgb(229, 27, 222)  / dark pink table header rgb(180, 0, 160)
  • Watermark logo centred on every page
  • DejaVu font (Unicode) so "Rs." renders correctly — ₹ replaced with "Rs."
  • Section headings, stat cards, footer identical to settings_pdf.js

FIXES (v2):
  1. Content report  — added 'content' branch to _get_report_data_internal()
  2. Double "Rs. Rs." — removed hardcoded "Rs. " prefix from revenue strings;
     _safe() already converts ₹ → "Rs." so we now use the ₹ symbol as prefix
     and let _safe() handle the conversion uniformly.
  3. Column-header icons — added a COLUMN_ICONS map; draw_table() prepends a
     DejaVu-safe Unicode glyph to every column header automatically.
"""

from flask import Blueprint, jsonify, request, Response
from database.db import get_db_connection
from routes.admin.admin_auth import admin_required
from datetime import datetime, timedelta
import json, os, io, base64

analytics_bp = Blueprint('admin_analytics', __name__)


# ══════════════════════════════════════════════════════════════════════════════
#  EMOJI COLUMN-HEADER MAP  (real color emoji via NotoColorEmoji)
# ══════════════════════════════════════════════════════════════════════════════
# ── FONT AWESOME ICON SYSTEM ──────────────────────────────────────────────────
# Uses the same fa-solid-900.ttf that Font Awesome 6 loads in the browser,
# so PDF icons are identical to what you see on the analytics page.
# Icons are rasterized via Pillow → PNG → reportlab ImageReader.

FA_SOLID_FONT = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '../../fonts/fa-solid-900.ttf'
)

# ── EXACT FA6 codepoints — taken directly from the class names in analytics.html ──
# Each entry: 'fa-class-name': '\uXXXX'
FA_CP = {
    # used in KPI cards
    'fa-indian-rupee-sign': '\ue3bc',   # Revenue / Gross Revenue
    'fa-users':             '\uf0c0',   # Total Users
    'fa-bag-shopping':      '\uf290',   # Total Orders / Orders
    'fa-calendar-check':    '\uf274',   # Total Bookings
    'fa-coins':             '\uf51e',   # Total Commission
    'fa-arrow-up-right-dots':'\ue4b7',  # Online Orders Rev.
    'fa-truck':             '\uf0d1',   # COD Orders / COD Orders Rev.
    'fa-wallet':            '\uf555',   # Total Withdrawn / Balance
    'fa-user-check':        '\uf4fc',   # Active (7 days)
    'fa-user-plus':         '\uf234',   # New (30 days) / New Users
    'fa-lock':              '\uf023',   # Suspended
    'fa-images':            '\uf302',   # Total Posts
    'fa-heart':             '\uf004',   # Total Likes / Followers
    'fa-comment':           '\uf075',   # Total Comments
    'fa-store':             '\uf54e',   # Products Listed / Product
    'fa-globe':             '\uf0ac',   # Online Orders
    'fa-circle-check':      '\uf058',   # Completed
    'fa-clock':             '\uf017',   # Active / Pending
    'fa-percent':           '\uf295',   # Completion Rate
    # used in tab buttons
    'fa-chart-pie':         '\uf200',   # Overview tab
    'fa-photo-film':        '\ue222',   # Content tab
    'fa-trophy':            '\uf091',   # Leaderboard tab
    'fa-calendar-range':    '\ue0d6',   # Period label
    # used in table column headers
    'fa-chart-bar':         '\uf080',   # Total / Count columns
    'fa-chart-line':        '\uf201',   # Rate columns
    'fa-user':              '\uf007',   # Username / User
    'fa-envelope':          '\uf0e0',   # Email
    'fa-star':              '\uf005',   # Creator
    'fa-tag':               '\uf02b',   # Name / Type
    'fa-hashtag':           '\uf292',   # Count
    'fa-thumbtack':         '\uf08d',   # fallback
}

# ── KPI CARD ICONS — exact match to each kpi-card in analytics.html ──────────
# Key = substring to look for in the KPI label (lowercase)
# Value = FA class name → looked up in FA_CP above
_KPI_ICON_MAP = [
    # Overview tab
    ('revenue',     'fa-indian-rupee-sign'),
    ('commission',  'fa-coins'),
    ('gross',       'fa-indian-rupee-sign'),
    ('withdrawn',   'fa-wallet'),
    ('balance',     'fa-wallet'),
    ('online',      'fa-arrow-up-right-dots'),
    ('cod',         'fa-truck'),
    # Users tab
    ('total users', 'fa-users'),
    ('active',      'fa-user-check'),
    ('new',         'fa-user-plus'),
    ('signup',      'fa-user-plus'),
    ('suspended',   'fa-lock'),
    ('locked',      'fa-lock'),
    # Content tab
    ('posts',       'fa-images'),
    ('post',        'fa-images'),
    ('likes',       'fa-heart'),
    ('like',        'fa-heart'),
    ('follower',    'fa-heart'),
    ('comment',     'fa-comment'),
    ('product',     'fa-store'),
    ('listed',      'fa-store'),
    # Orders tab
    ('orders',      'fa-bag-shopping'),
    ('order',       'fa-bag-shopping'),
    # Bookings tab
    ('booking',     'fa-calendar-check'),
    ('completed',   'fa-circle-check'),
    ('pending',     'fa-clock'),
    ('rate',        'fa-percent'),
    # Misc
    ('total',       'fa-chart-bar'),
    ('count',       'fa-hashtag'),
    ('creator',     'fa-star'),
    ('user',        'fa-users'),
    ('earned',      'fa-coins'),
    ('amount',      'fa-indian-rupee-sign'),
]

# ── COLUMN HEADER ICONS — match analytics.html table column styling ───────────
_COL_ICON_MAP = [
    ('period',      'fa-calendar-range'),
    ('date',        'fa-calendar-range'),
    ('month',       'fa-calendar-range'),
    ('week',        'fa-calendar-range'),
    ('year',        'fa-calendar-range'),
    ('commission',  'fa-coins'),
    ('revenue',     'fa-indian-rupee-sign'),
    ('amount',      'fa-indian-rupee-sign'),
    ('gross',       'fa-indian-rupee-sign'),
    ('earned',      'fa-coins'),
    ('withdrawn',   'fa-wallet'),
    ('balance',     'fa-wallet'),
    ('available',   'fa-wallet'),
    ('pending',     'fa-clock'),
    ('transaction', 'fa-chart-line'),
    ('order',       'fa-bag-shopping'),
    ('booking',     'fa-calendar-check'),
    ('total',       'fa-chart-bar'),
    ('online',      'fa-globe'),
    ('cod',         'fa-truck'),
    ('completed',   'fa-circle-check'),
    ('username',    'fa-user'),
    ('user',        'fa-user'),
    ('signup',      'fa-user-plus'),
    ('name',        'fa-tag'),
    ('email',       'fa-envelope'),
    ('creator',     'fa-star'),
    ('follower',    'fa-heart'),
    ('like',        'fa-heart'),
    ('post',        'fa-images'),
    ('product',     'fa-store'),
    ('listed',      'fa-store'),
    ('comment',     'fa-comment'),
    ('type',        'fa-tag'),
    ('status',      'fa-circle-check'),
    ('rate',        'fa-percent'),
    ('count',       'fa-hashtag'),
    ('suspended',   'fa-lock'),
]

# Cache: (codepoint, size_pt, icon_rgb, bg_rgb) → PNG bytes
_FA_PNG_CACHE = {}

def _fa_png_bytes(codepoint, size_pt, icon_rgb=(255, 255, 255), bg_rgb=(180, 0, 160)):
    """
    Rasterize a Font Awesome glyph to PNG bytes.
    icon_rgb  : fill colour of the icon (white for table headers, brand-color for cards)
    bg_rgb    : background colour to composite onto (eliminates alpha blending issues)
    Cached — same glyph+size+colours only renders once.
    """
    key = (codepoint, size_pt, icon_rgb, bg_rgb)
    if key in _FA_PNG_CACHE:
        return _FA_PNG_CACHE[key]
    try:
        from PIL import Image, ImageDraw, ImageFont
        import io as _io
        render_px = max(int(size_pt * 3), 32)          # 3× oversampling for crispness
        font      = ImageFont.truetype(FA_SOLID_FONT, render_px)
        pad       = render_px // 4
        canvas_px = render_px + pad * 2
        img  = Image.new('RGBA', (canvas_px, canvas_px), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        draw.text((pad, pad), codepoint, font=font, fill=(*icon_rgb, 255))
        bbox = img.getbbox()
        if not bbox:
            _FA_PNG_CACHE[key] = None
            return None
        img = img.crop(bbox)
        # Composite onto solid background (reportlab can't handle alpha PNGs reliably)
        bg = Image.new('RGB', img.size, bg_rgb)
        bg.paste(img, mask=img.split()[3])
        buf = _io.BytesIO()
        bg.save(buf, format='PNG')
        png = buf.getvalue()
        _FA_PNG_CACHE[key] = png
        return png
    except Exception as e:
        print("Icon render error:", e)
        _FA_PNG_CACHE[key] = None
        return None


def _fa_reader(codepoint, size_pt, icon_rgb=(255, 255, 255), bg_rgb=(180, 0, 160)):
    """
    Return a FRESH ImageReader every call.
    (ImageReader wraps a BytesIO that is exhausted after reportlab reads it once —
    always create a new one from the cached PNG bytes.)
    Returns (ImageReader, width_pt, height_pt) or None.
    """
    png = _fa_png_bytes(codepoint, size_pt, icon_rgb, bg_rgb)
    if png is None:
        return None
    try:
        from reportlab.lib.utils import ImageReader
        import io as _io
        return ImageReader(_io.BytesIO(png)), size_pt, size_pt
    except Exception:
        return None


def _col_icon_cp(col_name):
    """Return the FA codepoint for a table column header — matches analytics.html."""
    lo = col_name.lower()
    for kw, fa_class in _COL_ICON_MAP:
        if kw in lo:
            return FA_CP.get(fa_class, FA_CP['fa-thumbtack'])
    return FA_CP['fa-thumbtack']


def _kpi_icon_cp(label):
    """Return the FA codepoint for a KPI stat card — matches analytics.html kpi-icon."""
    lo = label.lower()
    for kw, fa_class in _KPI_ICON_MAP:
        if kw in lo:
            return FA_CP.get(fa_class, FA_CP['fa-chart-bar'])
    return FA_CP['fa-chart-bar']


# ══════════════════════════════════════════════════════════════════════════════
#  PDF GENERATION
# ══════════════════════════════════════════════════════════════════════════════

def _get_logo_bytes():
    _this_dir = os.path.dirname(os.path.abspath(__file__)) if '__file__' in dir() else os.getcwd()
    candidates = [
        os.path.join(_this_dir, '..', '..', 'frontend', 'images', 'logo2.png'),
        os.path.join(_this_dir, '..', '..', 'images', 'logo2.png'),
        os.path.join(_this_dir, '..', 'frontend', 'images', 'logo2.png'),
        os.path.join(_this_dir, '..', 'images', 'logo2.png'),
        'frontend/images/logo2.png',
        'images/logo2.png',
    ]
    for p in candidates:
        p = os.path.normpath(p)
        if os.path.isfile(p):
            with open(p, 'rb') as f:
                return f.read()
    return None


def _safe(text):
    """
    Strip / replace characters outside Latin-1 so Helvetica never corrupts.
    FIX: ₹ → 'Rs.' (single conversion; callers must NOT add 'Rs.' prefix
    themselves — just use the ₹ symbol and let _safe() do the work).
    """
    if text is None:
        return ''
    s = str(text)
    replacements = {
        '\u2018': "'", '\u2019': "'", '\u02BC': "'", '\u02B9': "'",
        '\u201C': '"', '\u201D': '"', '\u00AB': '"', '\u00BB': '"',
        '\u2013': '-', '\u2012': '-',
        '\u2014': '--',
        '\u2026': '...',
        '\u20B9': 'Rs.', '\u20A8': 'Rs.',   # ₹ / ₨ → Rs.
        '\u00A0': ' ',
        '\u2022': '-', '\u25CF': '-', '\u25AA': '-', '\u2023': '-',
        # DejaVu-safe glyphs used as icons — keep them as-is
        # (they are already in the DejaVu repertoire)
    }
    for bad, good in replacements.items():
        s = s.replace(bad, good)
    # Strip supplementary-plane characters (U+10000+) that DejaVu cannot render.
    # NOTE: emoji are drawn as images via Pillow, never passed through _safe(),
    # so stripping them here is safe and intentional.
    s = ''.join(c if ord(c) <= 0xFFFF else '' for c in s)
    return s.strip()


def _fmt_rs(value):
    """
    Format a numeric value as a rupee string using the ₹ symbol.
    _safe() will convert ₹ → 'Rs.' when writing to PDF, so callers
    should ALWAYS use this helper instead of building 'Rs. ...' manually.
    This prevents the double 'Rs. Rs.' bug.
    """
    return f'\u20B9{float(value):,.2f}'   # ₹12,345.67

def _generate_pdf(report_type, date_from, date_to, kpis, rows, columns):
    """
    Clean, simple PDF export:
    - White background
    - Single pink header bar with title
    - KPI summary as plain text lines (no cards)
    - Data table with pink header row
    - Simple footer with page number
    """
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError as e:
        raise RuntimeError(f"reportlab not installed: {e}")

    # ── Fonts ─────────────────────────────────────────────────────────────────
    FONT_DIR = '/usr/share/fonts/truetype/dejavu/'
    try:
        pdfmetrics.registerFont(TTFont('DJV',   FONT_DIR + 'DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DJV-B', FONT_DIR + 'DejaVuSans-Bold.ttf'))
        FONT      = 'DJV'
        FONT_BOLD = 'DJV-B'
    except Exception:
        FONT = FONT_BOLD = 'Helvetica'

    # ── Page constants ────────────────────────────────────────────────────────
    PW, PH = A4
    MX     = 14 / 25.4 * 72      # 14 mm margin
    CW     = PW - 2 * MX
    def mm(v): return v / 25.4 * 72

    # ── Brand colours ─────────────────────────────────────────────────────────
    PINK       = (229, 27, 222)
    PINK_LIGHT = (253, 236, 254)
    WHITE      = (255, 255, 255)
    DARK       = (28, 14, 40)
    GRAY       = (120, 110, 130)
    PINK_HDR   = (180, 0, 160)    # darker pink for table header

    def rgb(t): return t[0]/255, t[1]/255, t[2]/255

    report_labels = {
        'overview': 'Platform Overview',
        'revenue':  'Revenue & Commissions',
        'users':    'User Activity',
        'orders':   'Orders Summary',
        'bookings': 'Booking Analysis',
        'creators': 'Creator Performance',
        'content':  'Content Analytics',
    }
    report_label = report_labels.get(report_type, report_type.title())

    buf    = io.BytesIO()
    c      = rl_canvas.Canvas(buf, pagesize=A4)
    page_n = [1]

    # ── Header ────────────────────────────────────────────────────────────────
    def draw_header():
        hH = mm(14)
        c.setFillColorRGB(*rgb(PINK))
        c.rect(0, PH - hH, PW, hH, fill=1, stroke=0)

        c.setFillColorRGB(*rgb(WHITE))
        c.setFont(FONT_BOLD, 13)
        c.drawString(MX, PH - hH + mm(4), _safe(f'CreatorConnect \u2014 {report_label} Report'))

        ts = datetime.now().strftime('%d %b %Y, %H:%M')
        c.setFont(FONT, 7)
        c.setFillColorRGB(1, 0.9, 1)
        c.drawRightString(PW - MX, PH - hH + mm(4), _safe(f'Period: {date_from}  to  {date_to}    Generated: {ts}'))

        return PH - hH - mm(6)   # y after header

    # ── Footer ────────────────────────────────────────────────────────────────
    def draw_footer():
        fH = mm(9)
        c.setFillColorRGB(0.97, 0.94, 0.99)
        c.rect(0, 0, PW, fH, fill=1, stroke=0)
        c.setStrokeColorRGB(*rgb(PINK))
        c.setLineWidth(0.4)
        c.line(0, fH, PW, fH)
        c.setFont(FONT, 6.5)
        c.setFillColorRGB(*rgb(GRAY))
        c.drawString(MX, mm(2.5), 'CreatorConnect \u2014 Confidential')
        c.drawRightString(PW - MX, mm(2.5), f'Page {page_n[0]}')

    # ── New page helper ───────────────────────────────────────────────────────
    def new_page():
        draw_footer()
        c.showPage()
        page_n[0] += 1
        return draw_header()

    # ── KPI lines ─────────────────────────────────────────────────────────────
    def draw_kpis(y, kpis_dict):
        if not kpis_dict:
            return y
        if y - mm(8) < mm(20):
            y = new_page()

        c.setFillColorRGB(*rgb(PINK_LIGHT))
        block_h = mm(7) * len(kpis_dict) + mm(6)
        c.roundRect(MX, y - block_h, CW, block_h, mm(2), fill=1, stroke=0)

        cy = y - mm(5)
        for label, val in kpis_dict.items():
            c.setFont(FONT_BOLD, 8)
            c.setFillColorRGB(*rgb(PINK_HDR))
            c.drawString(MX + mm(4), cy, _safe(str(label)) + ':')
            label_w = c.stringWidth(_safe(str(label)) + ':', FONT_BOLD, 8)
            c.setFont(FONT, 8)
            c.setFillColorRGB(*rgb(DARK))
            c.drawString(MX + mm(4) + label_w + 4, cy, _safe(str(val)))
            cy -= mm(7)

        return y - block_h - mm(6)

    # ── Table ─────────────────────────────────────────────────────────────────
    def draw_table(y, col_headers, rows_data):
        if not col_headers or not rows_data:
            return y

        n        = len(col_headers)
        col_w    = CW / n
        row_h    = mm(7)
        hdr_h    = mm(9)

        def draw_hdr_row(yy):
            c.setFillColorRGB(*rgb(PINK_HDR))
            c.rect(MX, yy - hdr_h, CW, hdr_h, fill=1, stroke=0)
            c.setFont(FONT_BOLD, 7.5)
            c.setFillColorRGB(1, 1, 1)
            for i, hdr in enumerate(col_headers):
                cx = MX + i * col_w
                c.drawString(cx + mm(2), yy - hdr_h + mm(2.5), _safe(str(hdr)))
            return yy - hdr_h

        if y - hdr_h < mm(20):
            y = new_page()

        y = draw_hdr_row(y)

        for r_idx, row in enumerate(rows_data):
            if y - row_h < mm(14):
                y = new_page()
                y = draw_hdr_row(y)

            # Alternating row background
            if r_idx % 2 == 1:
                c.setFillColorRGB(*rgb(PINK_LIGHT))
                c.rect(MX, y - row_h, CW, row_h, fill=1, stroke=0)

            c.setFont(FONT, 7.5)
            c.setFillColorRGB(*rgb(DARK))
            for i, hdr in enumerate(col_headers):
                val = row.get(hdr, '--') if isinstance(row, dict) else (row[i] if i < len(row) else '--')
                txt = _safe(str(val)) if val is not None else '--'
                max_chars = int(col_w / mm(2.1))
                if len(txt) > max_chars:
                    txt = txt[:max_chars - 1] + '.'
                cx = MX + i * col_w
                c.drawString(cx + mm(2), y - row_h + mm(2.2), txt)

            # Row divider
            c.setStrokeColorRGB(0.90, 0.86, 0.95)
            c.setLineWidth(0.25)
            c.line(MX, y - row_h, MX + CW, y - row_h)
            y -= row_h

        # Outer border
        c.setStrokeColorRGB(*rgb(PINK))
        c.setLineWidth(0.5)
        total_rows = len(rows_data)
        table_h = hdr_h + total_rows * row_h
        c.rect(MX, y, CW, table_h, stroke=1, fill=0)

        return y - mm(8)

    # ── BUILD ─────────────────────────────────────────────────────────────────
    y = draw_header()

    # KPI summary block
    if kpis:
        y = draw_kpis(y, kpis)

    # Data table
    if columns and rows:
        # Section label
        if y - mm(8) < mm(20):
            y = new_page()
        c.setFont(FONT_BOLD, 8)
        c.setFillColorRGB(*rgb(PINK))
        c.drawString(MX, y, _safe(f'{report_label} Data  ({len(rows)} records)'))
        y -= mm(5)

        y = draw_table(y, columns, rows)

    draw_footer()
    c.save()
    buf.seek(0)
    return buf.read()
# ══════════════════════════════════════════════════════════════════════════════
#  HELPER UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def _date_params():
    today     = datetime.now().date()
    date_from = request.args.get('from', str(today - timedelta(days=30)))
    date_to   = request.args.get('to',   str(today))
    return date_from, date_to


def _iso(v):
    if v is None: return None
    if isinstance(v, datetime): return v.isoformat()
    if hasattr(v, 'isoformat'): return v.isoformat()
    return v


def _get_report_data_internal(report_type, df, dt, group):
    group_fmt = {
        'day':   '%Y-%m-%d',
        'week':  '%Y-%u',
        'month': '%Y-%m',
        'year':  '%Y',
    }.get(group, '%Y-%m')

    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)
    kpis    = {}
    rows    = []
    columns = []

    try:
        if report_type == 'overview':
            cur.execute("SELECT COUNT(*) AS v FROM users WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['New Users'] = int(cur.fetchone()['v'])
            cur.execute("SELECT COALESCE(SUM(commission_amt),0) AS v FROM commission_ledger WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Commission Earned'] = _fmt_rs(cur.fetchone()['v'])
            cur.execute("SELECT COUNT(*) AS v FROM product_orders WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Orders'] = int(cur.fetchone()['v'])
            cur.execute("SELECT COUNT(*) AS v FROM service_bookings WHERE booking_date BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Bookings'] = int(cur.fetchone()['v'])

            cur.execute(f"""
                SELECT DATE_FORMAT(u.created_at, '{group_fmt}') AS period,
                    COUNT(*) AS new_users
                FROM users u WHERE u.created_at BETWEEN %s AND %s
                GROUP BY period ORDER BY period
            """, (df, dt + ' 23:59:59'))
            signups_by_period = {r['period']: r['new_users'] for r in cur.fetchall()}

            cur.execute(f"""
                SELECT DATE_FORMAT(po.created_at, '{group_fmt}') AS period,
                    COUNT(*) AS orders
                FROM product_orders po WHERE po.created_at BETWEEN %s AND %s
                GROUP BY period ORDER BY period
            """, (df, dt + ' 23:59:59'))
            orders_by_period = {r['period']: r['orders'] for r in cur.fetchall()}

            all_periods = sorted(set(list(signups_by_period.keys()) + list(orders_by_period.keys())))
            rows = [{'Period': p, 'New Users': signups_by_period.get(p, 0), 'Orders': orders_by_period.get(p, 0)} for p in all_periods]
            columns = ['Period', 'New Users', 'Orders']

        elif report_type == 'revenue':
            cur.execute("SELECT COALESCE(SUM(commission_amt),0) AS v FROM commission_ledger WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Total Commission'] = _fmt_rs(cur.fetchone()['v'])
            cur.execute("SELECT COALESCE(SUM(amount),0) AS v FROM withdrawal_requests WHERE status IN ('approved','completed') AND processed_date BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Total Withdrawn'] = _fmt_rs(cur.fetchone()['v'])
            cur.execute("SELECT COUNT(*) AS v FROM withdrawal_requests WHERE status='pending'")
            kpis['Pending Withdrawals'] = int(cur.fetchone()['v'])
            cur.execute(f"""
                SELECT DATE_FORMAT(created_at, '{group_fmt}') AS period,
                       COALESCE(SUM(commission_amt),0) AS commission,
                       COALESCE(SUM(gross_amount),0) AS gross_amount,
                       COUNT(*) AS transactions
                FROM commission_ledger WHERE created_at BETWEEN %s AND %s GROUP BY period ORDER BY period
            """, (df, dt + ' 23:59:59'))
            rows = [{'Period': r['period'],
                     'Commission (Rs.)': _fmt_rs(r['commission']),
                     'Gross Amount (Rs.)': _fmt_rs(r['gross_amount']),
                     'Transactions': r['transactions']} for r in cur.fetchall()]
            columns = ['Period', 'Commission (Rs.)', 'Gross Amount (Rs.)', 'Transactions']

        elif report_type == 'users':
            cur.execute("SELECT COUNT(*) AS v FROM users WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['New Users'] = int(cur.fetchone()['v'])
            cur.execute("SELECT COUNT(*) AS v FROM users WHERE last_login BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Active Users'] = int(cur.fetchone()['v'])
            cur.execute("SELECT COUNT(*) AS v FROM users WHERE account_locked_until IS NOT NULL AND account_locked_until > NOW()")
            kpis['Suspended'] = int(cur.fetchone()['v'])
            cur.execute(f"""
                SELECT DATE_FORMAT(created_at, '{group_fmt}') AS period, COUNT(*) AS signups
                FROM users WHERE created_at BETWEEN %s AND %s GROUP BY period ORDER BY period
            """, (df, dt + ' 23:59:59'))
            rows = [{'Period': r['period'], 'New Signups': r['signups']} for r in cur.fetchall()]
            columns = ['Period', 'New Signups']

        elif report_type == 'orders':
            cur.execute("SELECT COUNT(*) AS v FROM product_orders WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Total Orders'] = int(cur.fetchone()['v'])
            cur.execute("SELECT COALESCE(SUM(total_amount),0) AS v FROM product_orders WHERE status NOT IN ('cancelled','returned') AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Gross Revenue'] = _fmt_rs(cur.fetchone()['v'])
            cur.execute("SELECT COUNT(*) AS v FROM product_orders WHERE payment_method='cod' AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['COD Orders'] = int(cur.fetchone()['v'])
            cur.execute(f"""
                SELECT DATE_FORMAT(created_at, '{group_fmt}') AS period, COUNT(*) AS orders,
                       SUM(CASE WHEN payment_method='cod' THEN 1 ELSE 0 END) AS cod,
                       SUM(CASE WHEN payment_method!='cod' THEN 1 ELSE 0 END) AS online,
                       COALESCE(SUM(total_amount),0) AS revenue
                FROM product_orders WHERE created_at BETWEEN %s AND %s GROUP BY period ORDER BY period
            """, (df, dt + ' 23:59:59'))
            rows = [{'Period': r['period'],
                     'Total Orders': r['orders'],
                     'Online': r['online'],
                     'COD': r['cod'],
                     'Revenue (Rs.)': _fmt_rs(r['revenue'])} for r in cur.fetchall()]
            columns = ['Period', 'Total Orders', 'Online', 'COD', 'Revenue (Rs.)']

        elif report_type == 'bookings':
            cur.execute("SELECT COUNT(*) AS v FROM service_bookings WHERE booking_date BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            total = int(cur.fetchone()['v'])
            kpis['Total Bookings'] = total
            cur.execute("SELECT COUNT(*) AS v FROM service_bookings WHERE status='completed' AND booking_date BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            completed = int(cur.fetchone()['v'])
            kpis['Completed'] = completed
            kpis['Completion Rate'] = f"{(completed/total*100):.1f}%" if total else "0%"
            cur.execute(f"""
                SELECT DATE_FORMAT(booking_date, '{group_fmt}') AS period, COUNT(*) AS bookings,
                       SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) AS completed,
                       COALESCE(SUM(CASE WHEN status='completed' THEN total_amount ELSE 0 END),0) AS revenue
                FROM service_bookings WHERE booking_date BETWEEN %s AND %s GROUP BY period ORDER BY period
            """, (df, dt + ' 23:59:59'))
            rows = [{'Period': r['period'],
                     'Total': r['bookings'],
                     'Completed': r['completed'],
                     'Revenue (Rs.)': _fmt_rs(r['revenue'])} for r in cur.fetchall()]
            columns = ['Period', 'Total', 'Completed', 'Revenue (Rs.)']

        elif report_type == 'creators':
            cur.execute("""
                SELECT u.username, u.full_name, u.email,
                       sb.available_balance, sb.total_earnings, sb.total_withdrawn,
                       (SELECT COUNT(*) FROM posts p WHERE p.user_id = u.id AND p.is_deleted=0) AS post_count,
                       (SELECT COUNT(*) FROM followers f WHERE f.following_id = u.id) AS followers
                FROM seller_balance sb JOIN users u ON sb.user_id = u.id
                ORDER BY sb.total_earnings DESC LIMIT 50
            """)
            db_rows = cur.fetchall()
            rows = [{
                'Username':           r['username'],
                'Full Name':          r['full_name'] or '--',
                'Total Earned (Rs.)': _fmt_rs(r['total_earnings']),
                'Available (Rs.)':    _fmt_rs(r['available_balance']),
                'Withdrawn (Rs.)':    _fmt_rs(r['total_withdrawn']),
                'Posts':              r['post_count'],
                'Followers':          r['followers'],
            } for r in db_rows]
            columns = ['Username', 'Full Name', 'Total Earned (Rs.)', 'Available (Rs.)', 'Withdrawn (Rs.)', 'Posts', 'Followers']
            kpis['Creators Listed'] = len(rows)

        # ── FIX: Content report branch (was completely missing) ───────────────
        elif report_type == 'content':
            cur.execute("SELECT COUNT(*) AS v FROM posts WHERE is_deleted = 0 AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Total Posts'] = int(cur.fetchone()['v'])

            cur.execute("SELECT COALESCE(SUM(likes_count),0) AS v FROM posts WHERE is_deleted = 0 AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Total Likes'] = int(cur.fetchone()['v'])

            try:
                cur.execute("SELECT COUNT(*) AS v FROM comments WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
                kpis['Total Comments'] = int(cur.fetchone()['v'])
            except Exception:
                kpis['Total Comments'] = 0

            cur.execute("SELECT COUNT(*) AS v FROM posts WHERE is_deleted = 0 AND post_type IN ('product','service') AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
            kpis['Products Listed'] = int(cur.fetchone()['v'])

            cur.execute(f"""
                SELECT DATE_FORMAT(created_at, '{group_fmt}') AS period,
                       COUNT(*) AS post_count,
                       COALESCE(SUM(likes_count), 0) AS likes,
                       COALESCE(SUM(CASE WHEN post_type IN ('product','service') THEN 1 ELSE 0 END), 0) AS products
                FROM posts
                WHERE is_deleted = 0 AND created_at BETWEEN %s AND %s
                GROUP BY period ORDER BY period
            """, (df, dt + ' 23:59:59'))
            rows = [{
                'Period':    r['period'],
                'Posts':     r['post_count'],
                'Likes':     r['likes'],
                'Products':  r['products'],
            } for r in cur.fetchall()]
            columns = ['Period', 'Posts', 'Likes', 'Products']

    finally:
        cur.close()
        conn.close()

    return {'kpis': kpis, 'rows': rows, 'columns': columns}


# ══════════════════════════════════════════════════════════════════════════════
#  OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════

@analytics_bp.route('/api/admin/analytics/overview', methods=['GET'])
@admin_required
def overview():
    df, dt = _date_params()
    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)

    cur.execute("SELECT COALESCE(SUM(commission_amt), 0) AS v FROM commission_ledger WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    total_commission = float(cur.fetchone()['v'])

    cur.execute("SELECT COUNT(*) AS v FROM users")
    total_users = int(cur.fetchone()['v'])

    cur.execute("SELECT COUNT(*) AS v FROM product_orders WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    total_orders = int(cur.fetchone()['v'])

    cur.execute("SELECT COUNT(*) AS v FROM product_orders WHERE payment_method != 'cod' AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    online_orders = int(cur.fetchone()['v'])
    cod_orders = total_orders - online_orders

    cur.execute("SELECT COUNT(*) AS v FROM service_bookings WHERE booking_date BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    total_bookings = int(cur.fetchone()['v'])

    cur.execute("""
        SELECT DATE_FORMAT(created_at, '%Y-%m-%d') AS day, COALESCE(SUM(commission_amt), 0) AS commission
        FROM commission_ledger
        WHERE created_at BETWEEN %s AND %s
        GROUP BY day ORDER BY day ASC
    """, (df, dt + ' 23:59:59'))
    monthly_commission = [{'month': r['day'], 'commission': float(r['commission'])} for r in cur.fetchall()]

    cur.execute("""
        SELECT DATE_FORMAT(created_at, '%Y-%m-%d') AS day, COUNT(*) AS cnt
        FROM users WHERE created_at BETWEEN %s AND %s
        GROUP BY day ORDER BY day ASC
    """, (df, dt + ' 23:59:59'))
    monthly_users = [{'month': r['day'], 'cnt': int(r['cnt'])} for r in cur.fetchall()]

    cur.execute("""
        SELECT DATE_FORMAT(created_at, '%Y-%m-%d') AS day, COUNT(*) AS cnt
        FROM posts WHERE is_deleted = 0 AND created_at BETWEEN %s AND %s
        GROUP BY day ORDER BY day ASC
    """, (df, dt + ' 23:59:59'))
    monthly_posts = [{'month': r['day'], 'cnt': int(r['cnt'])} for r in cur.fetchall()]

    cur.close(); conn.close()
    return jsonify({
        'total_commission': total_commission,
        'total_users': total_users,
        'total_orders': total_orders,
        'online_orders': online_orders,
        'cod_orders': cod_orders,
        'total_bookings': total_bookings,
        'monthly_commission': monthly_commission,
        'monthly_users': monthly_users,
        'monthly_posts': monthly_posts,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  REVENUE
# ══════════════════════════════════════════════════════════════════════════════

@analytics_bp.route('/api/admin/analytics/revenue', methods=['GET'])
@admin_required
def revenue():
    df, dt = _date_params()
    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)

    cur.execute("SELECT COALESCE(SUM(commission_amt), 0) AS v FROM commission_ledger WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    total_commission = float(cur.fetchone()['v'])

    cur.execute("SELECT COALESCE(SUM(commission_amt), 0) AS v FROM commission_ledger WHERE event_type = 'online_commission' AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    online_commission = float(cur.fetchone()['v'])

    cur.execute("SELECT COALESCE(SUM(commission_amt), 0) AS v FROM commission_ledger WHERE event_type = 'cod_commission' AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    cod_commission = float(cur.fetchone()['v'])

    cur.execute("SELECT COALESCE(SUM(commission_amt), 0) AS v FROM commission_ledger WHERE event_type = 'cod_deficit' AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    cod_deficit = float(cur.fetchone()['v'])

    cur.execute("SELECT COALESCE(SUM(commission_amt), 0) AS v FROM commission_ledger WHERE event_type = 'deficit_recovery' AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    deficit_recovery = float(cur.fetchone()['v'])

    cur.execute("SELECT COALESCE(SUM(amount), 0) AS v FROM withdrawal_requests WHERE status IN ('approved','completed') AND processed_date BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    total_withdrawn = float(cur.fetchone()['v'])

    cur.execute("SELECT COALESCE(SUM(amount), 0) AS v FROM withdrawal_requests WHERE status = 'pending'")
    wd_pending = float(cur.fetchone()['v'])
    cur.execute("SELECT COALESCE(SUM(amount), 0) AS v FROM withdrawal_requests WHERE status IN ('approved','completed')")
    wd_approved = float(cur.fetchone()['v'])
    cur.execute("SELECT COALESCE(SUM(amount), 0) AS v FROM withdrawal_requests WHERE status = 'rejected'")
    wd_rejected = float(cur.fetchone()['v'])

    cur.execute("""
        SELECT DATE_FORMAT(created_at, '%Y-%m-%d') AS day, COALESCE(SUM(commission_amt), 0) AS commission
        FROM commission_ledger WHERE created_at BETWEEN %s AND %s
        GROUP BY day ORDER BY day ASC
    """, (df, dt + ' 23:59:59'))
    monthly = [{'month': r['day'], 'commission': float(r['commission'])} for r in cur.fetchall()]

    cur.close(); conn.close()
    return jsonify({
        'total_commission': total_commission,
        'online_commission': online_commission,
        'cod_commission': cod_commission,
        'cod_deficit': cod_deficit,
        'deficit_recovery': deficit_recovery,
        'total_withdrawn': total_withdrawn,
        'wd_pending': wd_pending,
        'wd_approved': wd_approved,
        'wd_rejected': wd_rejected,
        'monthly': monthly,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  USERS
# ══════════════════════════════════════════════════════════════════════════════

@analytics_bp.route('/api/admin/analytics/users', methods=['GET'])
@admin_required
def users_analytics():
    df, dt = _date_params()
    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)

    cur.execute("SELECT COUNT(*) AS v FROM users")
    total_users = int(cur.fetchone()['v'])
    cur.execute("SELECT COUNT(*) AS v FROM users WHERE last_login >= DATE_SUB(NOW(), INTERVAL 7 DAY)")
    active_week = int(cur.fetchone()['v'])
    cur.execute("SELECT COUNT(*) AS v FROM users WHERE created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)")
    new_month = int(cur.fetchone()['v'])
    cur.execute("SELECT COUNT(*) AS v FROM users WHERE account_locked_until IS NOT NULL AND account_locked_until > NOW()")
    locked = int(cur.fetchone()['v'])
    cur.execute("SELECT COUNT(*) AS v FROM users WHERE role = 0")
    creators = int(cur.fetchone()['v'])
    cur.execute("SELECT COUNT(*) AS v FROM users WHERE role = 1")
    admins = int(cur.fetchone()['v'])

    cur.execute("""
        SELECT DATE_FORMAT(created_at, '%Y-%m-%d') AS day, COUNT(*) AS cnt
        FROM users WHERE created_at BETWEEN %s AND %s
        GROUP BY day ORDER BY day ASC
    """, (df, dt + ' 23:59:59'))
    monthly_signups = [{'month': r['day'], 'cnt': int(r['cnt'])} for r in cur.fetchall()]

    cur.execute("""
        SELECT DATE_FORMAT(last_login, '%Y-%m-%d') AS day, COUNT(*) AS cnt
        FROM users WHERE last_login BETWEEN %s AND %s
        GROUP BY day ORDER BY day ASC
    """, (df, dt + ' 23:59:59'))
    active_trend = [{'month': r['day'], 'cnt': int(r['cnt'])} for r in cur.fetchall()]

    cur.execute("SELECT COALESCE(gender,'Not set') AS gender, COUNT(*) AS cnt FROM users GROUP BY gender")
    gender_dist = {r['gender']: int(r['cnt']) for r in cur.fetchall()}

    cur.close(); conn.close()
    return jsonify({
        'total_users': total_users,
        'active_week': active_week,
        'new_month': new_month,
        'locked': locked,
        'creators': creators,
        'admins': admins,
        'monthly_signups': monthly_signups,
        'active_trend': active_trend,
        'gender_dist': gender_dist,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  CONTENT
# ══════════════════════════════════════════════════════════════════════════════

@analytics_bp.route('/api/admin/analytics/content', methods=['GET'])
@admin_required
def content_analytics():
    df, dt = _date_params()
    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)

    cur.execute("SELECT COUNT(*) AS v FROM posts WHERE is_deleted = 0")
    total_posts = int(cur.fetchone()['v'])

    cur.execute("SELECT COALESCE(SUM(likes_count),0) AS v FROM posts WHERE is_deleted = 0")
    total_likes = int(cur.fetchone()['v'])

    try:
        cur.execute("SELECT COUNT(*) AS v FROM comments")
        total_comments = int(cur.fetchone()['v'])
    except Exception:
        total_comments = 0

    cur.execute("SELECT COUNT(*) AS v FROM posts WHERE is_deleted = 0 AND post_type IN ('product','service')")
    total_products = int(cur.fetchone()['v'])

    cur.execute("""
        SELECT DATE_FORMAT(created_at, '%Y-%m-%d') AS day, COUNT(*) AS cnt
        FROM posts WHERE is_deleted = 0 AND created_at BETWEEN %s AND %s
        GROUP BY day ORDER BY day ASC
    """, (df, dt + ' 23:59:59'))
    monthly_posts = [{'month': r['day'], 'cnt': int(r['cnt'])} for r in cur.fetchall()]

    cur.execute("""
        SELECT COALESCE(post_type,'post') AS post_type, COUNT(*) AS cnt
        FROM posts WHERE is_deleted = 0
        GROUP BY post_type
    """)
    by_type = {r['post_type']: int(r['cnt']) for r in cur.fetchall()}

    cur.close(); conn.close()
    return jsonify({
        'total_posts': total_posts,
        'total_likes': total_likes,
        'total_comments': total_comments,
        'total_products': total_products,
        'monthly_posts': monthly_posts,
        'by_type': by_type,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  ORDERS
# ══════════════════════════════════════════════════════════════════════════════

@analytics_bp.route('/api/admin/analytics/orders', methods=['GET'])
@admin_required
def orders_analytics():
    df, dt = _date_params()
    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)

    cur.execute("SELECT COUNT(*) AS v FROM product_orders WHERE created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    total_orders = int(cur.fetchone()['v'])

    cur.execute("SELECT COUNT(*) AS v FROM product_orders WHERE payment_method != 'cod' AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    online_orders = int(cur.fetchone()['v'])
    cod_orders = total_orders - online_orders

    cur.execute("SELECT COALESCE(SUM(total_amount),0) AS v FROM product_orders WHERE status NOT IN ('cancelled','returned') AND created_at BETWEEN %s AND %s", (df, dt + ' 23:59:59'))
    total_revenue = float(cur.fetchone()['v'])

    cur.execute("""
        SELECT
            DATE_FORMAT(created_at, '%Y-%m-%d') AS day,
            COUNT(*) AS total_count,
            SUM(CASE WHEN payment_method = 'cod' THEN 1 ELSE 0 END) AS cod_count,
            SUM(CASE WHEN payment_method != 'cod' THEN 1 ELSE 0 END) AS online_count,
            COALESCE(SUM(CASE WHEN status NOT IN ('cancelled','returned') THEN total_amount ELSE 0 END), 0) AS revenue
        FROM product_orders
        WHERE created_at BETWEEN %s AND %s
        GROUP BY day ORDER BY day ASC
    """, (df, dt + ' 23:59:59'))
    monthly = [{
        'month': r['day'],
        'total_count': int(r['total_count']),
        'cod_count': int(r['cod_count']),
        'online_count': int(r['online_count']),
        'revenue': float(r['revenue'])
    } for r in cur.fetchall()]

    cur.execute("""
        SELECT status, COUNT(*) AS cnt
        FROM product_orders WHERE created_at BETWEEN %s AND %s
        GROUP BY status
    """, (df, dt + ' 23:59:59'))
    by_status = {r['status']: int(r['cnt']) for r in cur.fetchall()}

    cur.close(); conn.close()
    return jsonify({
        'total_orders': total_orders,
        'online_orders': online_orders,
        'cod_orders': cod_orders,
        'total_revenue': total_revenue,
        'monthly': monthly,
        'by_status': by_status,
    })


# ══════════════════════════════════════════════════════════════════════════════
#  BOOKINGS
# ══════════════════════════════════════════════════════════════════════════════

@analytics_bp.route('/api/admin/analytics/bookings', methods=['GET'])
@admin_required
def bookings_analytics():
    df, dt = _date_params()
    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)

    cur.execute("SELECT status, COUNT(*) AS cnt FROM service_bookings WHERE booking_date BETWEEN %s AND %s GROUP BY status", (df, dt + ' 23:59:59'))
    by_status = {r['status']: int(r['cnt']) for r in cur.fetchall()}

    cur.execute("""
        SELECT DATE_FORMAT(booking_date, '%Y-%m-%d') AS day, COUNT(*) AS cnt,
               COALESCE(SUM(CASE WHEN status='completed' THEN total_amount ELSE 0 END), 0) AS revenue
        FROM service_bookings WHERE booking_date BETWEEN %s AND %s
        GROUP BY day ORDER BY day ASC
    """, (df, dt + ' 23:59:59'))
    monthly = [{'month': r['day'], 'cnt': int(r['cnt']), 'revenue': float(r['revenue'])} for r in cur.fetchall()]

    cur.close(); conn.close()
    return jsonify({'by_status': by_status, 'monthly': monthly})


# ══════════════════════════════════════════════════════════════════════════════
#  LEADERBOARD
# ══════════════════════════════════════════════════════════════════════════════

@analytics_bp.route('/api/admin/analytics/leaderboard', methods=['GET'])
@admin_required
def leaderboard():
    conn = get_db_connection()
    cur  = conn.cursor(dictionary=True)

    cur.execute("""
        SELECT u.id, u.username, u.full_name, u.profile_pic,
               sb.available_balance, sb.total_earnings,
               (SELECT COUNT(*) FROM posts p WHERE p.user_id = u.id AND p.is_deleted = 0) AS post_count
        FROM seller_balance sb
        JOIN users u ON sb.user_id = u.id
        ORDER BY sb.available_balance DESC LIMIT 10
    """)
    top_creators = cur.fetchall()

    cur.execute("""
        SELECT p.post_id, COALESCE(p.product_title, p.title) AS product_title,
               p.post_type, p.likes_count, u.username, u.full_name
        FROM posts p JOIN users u ON p.user_id = u.id
        WHERE p.is_deleted = 0
        ORDER BY p.likes_count DESC LIMIT 10
    """)
    top_posts = cur.fetchall()

    cur.execute("""
        SELECT u.id, u.username, u.full_name, u.profile_pic,
               COUNT(po.order_id) AS order_count
        FROM product_orders po
        JOIN users u ON po.seller_id = u.id
        GROUP BY u.id ORDER BY order_count DESC LIMIT 10
    """)
    top_sellers = cur.fetchall()

    cur.execute("""
        SELECT u.id, u.username, u.full_name, u.profile_pic,
               COUNT(f.follower_id) AS follower_count
        FROM followers f JOIN users u ON f.following_id = u.id
        GROUP BY u.id ORDER BY follower_count DESC LIMIT 10
    """)
    top_followed = cur.fetchall()

    cur.close(); conn.close()

    def _ser(row):
        from decimal import Decimal
        return {k: float(v) if isinstance(v, Decimal) else v for k, v in row.items()}

    return jsonify({
        'top_creators': [_ser(r) for r in top_creators],
        'top_posts':    [_ser(r) for r in top_posts],
        'top_sellers':  [_ser(r) for r in top_sellers],
        'top_followed': [_ser(r) for r in top_followed],
    })


# ══════════════════════════════════════════════════════════════════════════════
#  REPORT DATA ENDPOINT
# ══════════════════════════════════════════════════════════════════════════════

@analytics_bp.route('/api/admin/analytics/report', methods=['GET'])
@admin_required
def report_data():
    df, dt = _date_params()
    report_type = request.args.get('type', 'overview')
    group       = request.args.get('group', 'month')
    data = _get_report_data_internal(report_type, df, dt, group)
    return jsonify(data)


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT — PDF & XLSX
# ══════════════════════════════════════════════════════════════════════════════

@analytics_bp.route('/api/admin/analytics/export', methods=['GET'])
@admin_required
def export_report():
    fmt         = request.args.get('format', 'xls')
    report_type = request.args.get('type', 'overview')
    df, dt      = _date_params()
    group       = request.args.get('group', 'month')

    try:
        data    = _get_report_data_internal(report_type, df, dt, group)
        columns = data.get('columns', [])
        rows    = data.get('rows', [])
        kpis    = data.get('kpis', {})

        # ── XLSX ─────────────────────────────────────────────────────────────
        if fmt == 'xls':
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            report_labels = {
                'overview': 'Platform Overview', 'revenue': 'Revenue & Commissions',
                'users': 'User Activity', 'orders': 'Orders Summary',
                'bookings': 'Booking Analysis', 'creators': 'Creator Performance',
                'content': 'Content Analytics',
            }
            ws.title = report_labels.get(report_type, report_type.title())[:31]

            BRAND_HEX = 'E60AEA'
            HDR_HEX   = 'B400A0'
            LIGHT_HEX = 'FAF5FF'
            ALT_HEX   = 'F5EEFF'

            hdr_fill  = PatternFill(start_color=HDR_HEX, end_color=HDR_HEX, fill_type='solid')
            hdr_font  = Font(color='FFFFFF', bold=True, size=10, name='Calibri')
            hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin      = Side(style='thin', color='E8D8F0')
            bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

            ws.merge_cells(f'A1:{get_column_letter(max(len(columns),1))}1')
            title_cell = ws['A1']
            title_cell.value = f'CreatorConnect — {ws.title} Report'
            title_cell.font  = Font(bold=True, size=14, color=BRAND_HEX, name='Calibri')
            title_cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[1].height = 24

            ws.merge_cells(f'A2:{get_column_letter(max(len(columns),1))}2')
            sub_cell = ws['A2']
            sub_cell.value = f'Period: {df} to {dt}   |   Generated: {datetime.now().strftime("%d %b %Y %H:%M")}'
            sub_cell.font  = Font(italic=True, size=9, color='6B5880', name='Calibri')
            sub_cell.alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 16

            if kpis:
                kpi_items = list(kpis.items())[:len(columns) or 4]
                kpi_col_end = get_column_letter(max(len(kpi_items), 1))
                ws.merge_cells(f'A3:{kpi_col_end}3')
                kpi_label_cell = ws['A3']
                kpi_label_cell.value = '  |  '.join([f'{k}: {v}' for k, v in kpi_items])
                kpi_label_cell.font  = Font(bold=True, size=9, color='7C3AED', name='Calibri')
                kpi_label_cell.alignment = Alignment(horizontal='center')
                kpi_label_cell.fill = PatternFill(start_color='F5EEFF', end_color='F5EEFF', fill_type='solid')
                ws.row_dimensions[3].height = 18
                start_row = 5
            else:
                start_row = 4

            if columns:
                for col_i, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=start_row, column=col_i, value=col_name)
                    cell.fill      = hdr_fill
                    cell.font      = hdr_font
                    cell.alignment = hdr_align
                    cell.border    = bdr
                ws.row_dimensions[start_row].height = 20

                for row_i, row in enumerate(rows, start_row + 1):
                    alt = row_i % 2 == 0
                    row_fill = PatternFill(start_color=ALT_HEX if alt else 'FFFFFF',
                                          end_color=ALT_HEX if alt else 'FFFFFF',
                                          fill_type='solid')
                    for col_i, col_name in enumerate(columns, 1):
                        val  = row.get(col_name, '')
                        cell = ws.cell(row=row_i, column=col_i, value=str(val))
                        cell.fill      = row_fill
                        cell.border    = bdr
                        cell.font      = Font(size=9, name='Calibri')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.row_dimensions[row_i].height = 16

                for col_i, col_name in enumerate(columns, 1):
                    col_letter = get_column_letter(col_i)
                    max_len = max(
                        len(str(col_name)),
                        *[len(str(row.get(col_name, ''))) for row in rows]
                    ) if rows else len(str(col_name))
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition':
                         f'attachment; filename=CreatorConnect_{report_type}_{df}_{dt}.xlsx'}
            )

        # ── PDF ───────────────────────────────────────────────────────────────
        elif fmt == 'pdf':
            pdf_bytes = _generate_pdf(
                report_type=report_type,
                date_from=df,
                date_to=dt,
                kpis=kpis,
                rows=rows,
                columns=columns,
            )
            return Response(
                pdf_bytes,
                mimetype='application/pdf',
                headers={'Content-Disposition':
                         f'attachment; filename=CreatorConnect_{report_type}_{df}_{dt}.pdf'}
            )

        else:
            return jsonify({'error': f'Unknown format: {fmt}'}), 400

    except ImportError as e:
        return jsonify({'error': f'Export library not installed: {e}. Run: pip install openpyxl reportlab Pillow pypdf'}), 501
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500